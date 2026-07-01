import datetime as dt
import os
import tempfile
import zipfile
from decimal import Decimal
from types import SimpleNamespace
from xml.etree import ElementTree as ET

from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from Usuarios.models import Usuario
from .models import (
    EjecucionConsolidacion,
    FacturaCorreo,
    FacturaRadian,
    ResultadoConciliacion,
)
from .utils.conciliacion import conciliar
from .utils.dian_client import DianClient
from .utils.export_xlsx import generar_excel
from .utils.xml_parser import parse_invoice_element, parse_zip


def _invoice_xml(numero="FE-001", cufe="CUFE-001", total="119.00"):
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<Invoice xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2"
         xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"
         xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2">
  <cbc:ID>{numero}</cbc:ID>
  <cbc:UUID>{cufe}</cbc:UUID>
  <cbc:IssueDate>2026-06-15</cbc:IssueDate>
  <cac:AccountingSupplierParty>
    <cac:Party>
      <cac:PartyTaxScheme>
        <cbc:RegistrationName>Proveedor SAS</cbc:RegistrationName>
        <cbc:CompanyID schemeName="NIT">900123456</cbc:CompanyID>
      </cac:PartyTaxScheme>
      <cac:Contact>
        <cbc:ElectronicMail>proveedor@example.com</cbc:ElectronicMail>
      </cac:Contact>
    </cac:Party>
  </cac:AccountingSupplierParty>
  <cac:AccountingCustomerParty>
    <cac:Party>
      <cac:PartyTaxScheme>
        <cbc:RegistrationName>Euro</cbc:RegistrationName>
        <cbc:CompanyID schemeName="NIT">800123456</cbc:CompanyID>
      </cac:PartyTaxScheme>
      <cac:Contact>
        <cbc:ElectronicMail>facturas@example.com</cbc:ElectronicMail>
      </cac:Contact>
    </cac:Party>
  </cac:AccountingCustomerParty>
  <cac:LegalMonetaryTotal>
    <cbc:LineExtensionAmount>100.00</cbc:LineExtensionAmount>
    <cbc:TaxInclusiveAmount>{total}</cbc:TaxInclusiveAmount>
    <cbc:PayableAmount>{total}</cbc:PayableAmount>
  </cac:LegalMonetaryTotal>
</Invoice>"""


class ConciliacionTests(TestCase):
    def test_concilia_por_cufe_y_deja_solo_correo_sin_match(self):
        ejecucion = SimpleNamespace(id=1)
        radian = [
            SimpleNamespace(
                id=1,
                cufe="CUFE-1",
                numero="F001",
                nit_proveedor="9001",
                nombre_proveedor="Proveedor",
                total=Decimal("100.00"),
                fecha_emision=dt.date(2026, 6, 15),
            )
        ]
        correos = [
            SimpleNamespace(
                id=10,
                cufe="CUFE-1",
                numero="OTRO",
                nit_proveedor="9001",
                nombre_proveedor="Proveedor",
                total=Decimal("100.00"),
                fecha_emision=dt.date(2026, 6, 15),
                fecha_correo=timezone.now(),
                asunto_correo="Factura conciliada",
            ),
            SimpleNamespace(
                id=11,
                cufe="SIN-MATCH",
                numero="F999",
                nit_proveedor="9002",
                nombre_proveedor="Otro",
                total=Decimal("50.00"),
                fecha_emision=dt.date(2026, 6, 15),
                fecha_correo=timezone.now(),
                asunto_correo="Solo correo",
            ),
        ]

        resultados = conciliar(ejecucion, radian, correos)

        self.assertEqual(resultados[0]["estado"], "CONCILIADA")
        self.assertEqual(resultados[0]["nivel_match"], "N1")
        self.assertEqual(resultados[1]["estado"], "SOLO_CORREO")

    def test_concilia_por_heuristica_n3_como_revision_manual(self):
        ejecucion = SimpleNamespace(id=1)
        radian = [
            SimpleNamespace(
                id=1,
                cufe="",
                numero="F001",
                nit_proveedor="9001",
                nombre_proveedor="Proveedor",
                total=Decimal("100.50"),
                fecha_emision=dt.date(2026, 6, 15),
            )
        ]
        correos = [
            SimpleNamespace(
                id=10,
                cufe="",
                numero="DIFERENTE",
                nit_proveedor="9001",
                nombre_proveedor="Proveedor",
                total=Decimal("100.00"),
                fecha_emision=dt.date(2026, 6, 17),
                fecha_correo=timezone.now(),
                asunto_correo="Factura revision",
            )
        ]

        resultados = conciliar(ejecucion, radian, correos)

        self.assertEqual(resultados[0]["estado"], "REVISION_MANUAL")
        self.assertEqual(resultados[0]["nivel_match"], "N3")


class XmlParserTests(TestCase):
    def test_parsea_invoice_ubl_directo(self):
        inv = ET.fromstring(_invoice_xml().encode("utf-8"))

        factura = parse_invoice_element(inv)

        self.assertEqual(factura.cufe, "CUFE-001")
        self.assertEqual(factura.numero, "FE-001")
        self.assertEqual(factura.emisor_nit, "900123456")
        self.assertEqual(factura.adquirente_correo, "facturas@example.com")
        self.assertEqual(factura.total, 119.00)

    def test_parsea_zip_con_xml(self):
        fd, zip_path = tempfile.mkstemp(suffix=".zip")
        os.close(fd)
        with zipfile.ZipFile(zip_path, "w") as zf:
            zf.writestr("factura.xml", _invoice_xml(numero="FE-002", cufe="CUFE-002"))

        try:
            factura = parse_zip(zip_path)
        finally:
            os.remove(zip_path)

        self.assertEqual(factura.numero, "FE-002")
        self.assertEqual(factura.cufe, "CUFE-002")


class ExportExcelTests(TestCase):
    def test_genera_excel_con_resumen_y_detalle(self):
        ejecucion = EjecucionConsolidacion.objects.create(
            fecha_desde=dt.date(2026, 6, 1),
            fecha_hasta=dt.date(2026, 6, 30),
            estado="COMPLETADA",
            total_radian=1,
            total_correo=1,
            total_conciliadas=1,
        )
        fr = FacturaRadian.objects.create(
            ejecucion=ejecucion,
            cufe="CUFE-1",
            numero="F001",
            nit_proveedor="9001",
            nombre_proveedor="Proveedor",
            total=Decimal("100.00"),
            fecha_emision=dt.date(2026, 6, 15),
            track_id="TRACK-1",
        )
        fc = FacturaCorreo.objects.create(
            ejecucion=ejecucion,
            cufe="CUFE-1",
            numero="F001",
            nit_proveedor="9001",
            nombre_proveedor="Proveedor",
            total=Decimal("100.00"),
            fecha_emision=dt.date(2026, 6, 15),
            asunto_correo="Factura",
            fecha_correo=timezone.now(),
            remitente="proveedor@example.com",
        )
        ResultadoConciliacion.objects.create(
            ejecucion=ejecucion,
            factura_radian=fr,
            factura_correo=fc,
            cufe="CUFE-1",
            numero="F001",
            nit_proveedor="9001",
            nombre_proveedor="Proveedor",
            monto_radian=Decimal("100.00"),
            monto_correo=Decimal("100.00"),
            delta_monto=Decimal("0.00"),
            fecha_radian=dt.date(2026, 6, 15),
            fecha_correo=timezone.now(),
            asunto_correo="Factura",
            estado="CONCILIADA",
            nivel_match="N1",
        )

        xlsx_bytes = generar_excel(ejecucion)
        workbook = load_workbook(filename=__import__("io").BytesIO(xlsx_bytes))

        self.assertEqual(workbook.sheetnames, ["Resumen", "Detalle"])
        self.assertEqual(
            workbook["Resumen"]["A1"].value,
            "Conciliaci\u00f3n RADIAN vs Correo \u2014 Resumen",
        )
        self.assertEqual(workbook["Detalle"]["A2"].value, "CONCILIADA")


class OptimizacionCorreosApiTests(TestCase):
    def setUp(self):
        self.user = Usuario.objects.create_user(
            correo="oc@example.com",
            nombres="Optimizacion",
            apellidos="Correos",
            cedula="100000001",
            password="test-pass-123",
        )
        permiso = Permission.objects.get(codename="can_view_optimizacion_correos")
        self.user.user_permissions.add(permiso)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def test_rechaza_rango_de_fechas_invalido(self):
        response = self.client.post(
            reverse("oc-ejecutar"),
            {"fecha_desde": "2026-06-10", "fecha_hasta": "2026-06-10"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["error"], "fecha_desde debe ser anterior a fecha_hasta")

    def test_rechaza_formato_de_fecha_invalido(self):
        response = self.client.post(
            reverse("oc-ejecutar"),
            {"fecha_desde": "2026/06/10", "fecha_hasta": "2026-06-11"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["error"], "Las fechas deben tener formato YYYY-MM-DD")


class DianClientSessionTests(TestCase):
    def test_detecta_html_de_login_como_sesion_no_autenticada(self):
        response = SimpleNamespace(
            url="https://catalogo-vpfe.dian.gov.co/User/CompanyLogin",
            text="<html><title>DIAN | Acceder</title></html>",
        )
        client = DianClient(nit="1", cc_rep="1")

        self.assertTrue(client._looks_like_login_response(response))
