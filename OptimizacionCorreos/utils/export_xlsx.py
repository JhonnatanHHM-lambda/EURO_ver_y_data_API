"""
Exportación Excel del resultado de conciliación con openpyxl.

Genera dos hojas:
  • Resumen — KPIs de la ejecución
  • Detalle — una fila por ResultadoConciliacion con colores por estado
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EURO_BLUE = "27348B"
EURO_BLUE_LIGHT = "EEF0F8"
WHITE = "FFFFFF"

ESTADO_COLORS = {
    "CONCILIADA":      "22C55E",
    "REVISION_MANUAL": "F59E0B",
    "SOLO_RADIAN":     "3B82F6",
    "SOLO_CORREO":     "F97316",
}

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _cell_font(bold=False, color=WHITE, size=11) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _fmt_date(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%d/%m/%Y")
    return str(v)


def _fmt_decimal(v) -> Optional[float]:
    if v is None:
        return None
    return float(Decimal(str(v)))


def generar_excel(ejecucion) -> bytes:
    """Genera el libro Excel y retorna sus bytes."""
    wb = Workbook()

    # ── Hoja Resumen ─────────────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumen"

    # Título
    ws_res.merge_cells("A1:B1")
    ws_res["A1"] = "Conciliación RADIAN vs Correo — Resumen"
    ws_res["A1"].font = Font(name="Calibri", bold=True, size=14, color=EURO_BLUE)
    ws_res["A1"].alignment = Alignment(horizontal="left")

    ws_res["A2"] = f"Rango: {_fmt_date(ejecucion.fecha_desde)} — {_fmt_date(ejecucion.fecha_hasta)}"
    ws_res["A2"].font = Font(name="Calibri", size=11, color="555555")

    ws_res.append([])

    kpis = [
        ("Total RADIAN", ejecucion.total_radian),
        ("Total Correo", ejecucion.total_correo),
        ("Conciliadas", ejecucion.total_conciliadas),
        ("Solo RADIAN", ejecucion.total_solo_radian),
        ("Solo Correo", ejecucion.total_solo_correo),
        ("Revisión Manual", ejecucion.total_revision),
    ]

    # Encabezado KPIs
    ws_res.append(["Indicador", "Valor"])
    for cell in ws_res[ws_res.max_row]:
        cell.fill = _header_fill(EURO_BLUE)
        cell.font = _cell_font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    for label, value in kpis:
        ws_res.append([label, value])
        row = ws_res.max_row
        ws_res.cell(row, 1).font = Font(name="Calibri", size=11)
        ws_res.cell(row, 1).border = BORDER
        ws_res.cell(row, 2).font = Font(name="Calibri", size=11, bold=True)
        ws_res.cell(row, 2).border = BORDER
        ws_res.cell(row, 2).alignment = Alignment(horizontal="center")
        ws_res.cell(row, 1).fill = PatternFill("solid", fgColor=EURO_BLUE_LIGHT)

    ws_res.column_dimensions["A"].width = 25
    ws_res.column_dimensions["B"].width = 15

    # ── Hoja Detalle ─────────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detalle")

    headers = [
        "Estado", "Nivel Match", "CUFE", "Número",
        "NIT Proveedor", "Nombre Proveedor",
        "Monto RADIAN", "Monto Correo", "Delta Monto",
        "Fecha RADIAN", "Fecha Correo", "Asunto Correo",
    ]
    ws_det.append(headers)
    for i, cell in enumerate(ws_det[1], 1):
        cell.fill = _header_fill(EURO_BLUE)
        cell.font = _cell_font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Anchos de columna
    col_widths = [18, 12, 45, 18, 18, 35, 16, 16, 14, 15, 18, 45]
    for i, w in enumerate(col_widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w
    ws_det.row_dimensions[1].height = 30

    resultados = ejecucion.resultados.select_related(
        "factura_radian", "factura_correo"
    ).all()

    for res in resultados:
        row = [
            res.estado,
            res.nivel_match or "",
            res.cufe or "",
            res.numero or "",
            res.nit_proveedor or "",
            res.nombre_proveedor or "",
            _fmt_decimal(res.monto_radian),
            _fmt_decimal(res.monto_correo),
            _fmt_decimal(res.delta_monto),
            _fmt_date(res.fecha_radian),
            _fmt_date(res.fecha_correo),
            res.asunto_correo or "",
        ]
        ws_det.append(row)
        row_idx = ws_det.max_row
        color_hex = ESTADO_COLORS.get(res.estado, "FFFFFF")
        estado_fill = PatternFill("solid", fgColor=color_hex)

        for col_idx in range(1, len(headers) + 1):
            cell = ws_det.cell(row_idx, col_idx)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3, 12)))
            if col_idx == 1:
                cell.fill = estado_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx in (7, 8, 9):
                cell.number_format = '#,##0.00'

    # Freeze header row
    ws_det.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
