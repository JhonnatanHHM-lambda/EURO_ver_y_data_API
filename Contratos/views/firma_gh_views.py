"""Vistas para el módulo de firma digital del GH (empleador)."""
import logging
from io import BytesIO

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

from EURO_ver_y_data.decoradores import require_permission
from Contratos.models import FirmaGH, FirmaProvisional, RegistroFirmaEmpleador

logger = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _usuario_info(u) -> dict:
    return {
        'id':      u.id,
        'nombres': u.nombres,
        'apellidos': u.apellidos,
        'nombre_completo': f'{u.nombres} {u.apellidos}'.strip(),
        'cedula':  u.cedula,
        'correo':  u.correo,
    }


def _estado_actual() -> dict:
    """Retorna el estado completo del módulo de firma."""
    firma_gh = FirmaGH.objects.select_related('usuario').first()
    firma_prov = FirmaProvisional.objects.select_related('usuario', 'autorizado_por').first()

    gh_data = None
    if firma_gh:
        gh_data = {
            'id':        firma_gh.id,
            'usuario':   _usuario_info(firma_gh.usuario),
            'habilitada': firma_gh.habilitada,
            'tiene_firma': bool(firma_gh.firma_imagen),
            'actualizado': firma_gh.actualizado.isoformat(),
        }

    prov_data = None
    if firma_prov:
        prov_data = {
            'id':          firma_prov.id,
            'usuario':     _usuario_info(firma_prov.usuario),
            'autorizado_por': _usuario_info(firma_prov.autorizado_por) if firma_prov.autorizado_por else None,
            'tiene_firma': bool(firma_prov.firma_imagen),
            'creado':      firma_prov.creado.isoformat(),
        }

    # Determina cuál firma está activa para los documentos
    if firma_gh and firma_gh.habilitada and firma_gh.firma_imagen:
        activa = 'gh'
    elif firma_prov and firma_prov.firma_imagen:
        activa = 'provisional'
    else:
        activa = None

    return {
        'firma_gh':          gh_data,
        'firma_provisional': prov_data,
        'activa':            activa,
    }


# ── Vistas ────────────────────────────────────────────────────────────────────

class FirmaGHView(APIView):
    """GET: estado actual. POST: crear/actualizar firma del GH."""
    permission_classes = [IsAuthenticated]

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def get(self, request):
        return Response(_estado_actual())

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def post(self, request):
        firma_imagen = request.data.get('firma_imagen', '').strip()
        if not firma_imagen:
            return Response({'error': 'firma_imagen es requerida.'}, status=status.HTTP_400_BAD_REQUEST)

        firma_gh, created = FirmaGH.objects.get_or_create(
            usuario=request.user,
            defaults={'firma_imagen': firma_imagen, 'habilitada': True},
        )
        if not created:
            firma_gh.firma_imagen = firma_imagen
            firma_gh.save(update_fields=['firma_imagen', 'actualizado'])

        return Response(_estado_actual(), status=status.HTTP_200_OK)


class FirmaGHToggleView(APIView):
    """POST: activa o desactiva la firma del GH."""
    permission_classes = [IsAuthenticated]

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def post(self, request):
        habilitar = request.data.get('habilitar')
        if habilitar is None:
            return Response({'error': 'El campo habilitar es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            firma_gh = FirmaGH.objects.get(usuario=request.user)
        except FirmaGH.DoesNotExist:
            return Response({'error': 'No tienes una firma registrada.'}, status=status.HTTP_404_NOT_FOUND)

        if not habilitar:
            # Solo puede desactivar si hay firma provisional activa
            if not FirmaProvisional.objects.exists():
                return Response(
                    {'error': 'Debes asignar una firma provisional antes de desactivar la tuya.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        firma_gh.habilitada = bool(habilitar)
        firma_gh.save(update_fields=['habilitada', 'actualizado'])

        # Al reactivar la firma del GH, eliminar la provisional
        if habilitar:
            FirmaProvisional.objects.all().delete()

        return Response(_estado_actual())


class FirmaProvisionaView(APIView):
    """POST: crear firma provisional. DELETE: eliminar firma provisional."""
    permission_classes = [IsAuthenticated]

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def post(self, request):
        usuario_id  = request.data.get('usuario_id')
        firma_imagen = request.data.get('firma_imagen', '').strip()

        if not usuario_id or not firma_imagen:
            return Response(
                {'error': 'usuario_id y firma_imagen son requeridos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            usuario_prov = User.objects.get(pk=usuario_id)
        except User.DoesNotExist:
            return Response({'error': 'Usuario no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        # Solo una provisional a la vez
        FirmaProvisional.objects.all().delete()
        FirmaProvisional.objects.create(
            usuario=usuario_prov,
            firma_imagen=firma_imagen,
            autorizado_por=request.user,
        )
        return Response(_estado_actual(), status=status.HTTP_201_CREATED)

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def delete(self, request):
        FirmaProvisional.objects.all().delete()
        return Response(_estado_actual())


class RegistroFirmaEmpleadorListView(APIView):
    """GET: lista el historial de documentos generados con firma del empleador."""
    permission_classes = [IsAuthenticated]

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def get(self, request):
        import pytz
        qs = (
            RegistroFirmaEmpleador.objects
            .select_related('contrato', 'contrato__sede', 'usuario_empleador')
            .order_by('-fecha_generacion')
        )

        fecha_desde = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta = request.query_params.get('fecha_hasta', '').strip()
        if fecha_desde:
            qs = qs.filter(fecha_generacion__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_generacion__date__lte=fecha_hasta)

        tz_co = pytz.timezone('America/Bogota')
        TIPO_CARTA = {'NO_PRORROGA': 'Sin prórroga', 'PRORROGA': 'Prórroga', 'TERMINACION': 'Terminación'}

        data = []
        for r in qs:
            c = r.contrato
            data.append({
                'id':                r.id,
                'fecha_generacion':  r.fecha_generacion.astimezone(tz_co).isoformat(),
                'tipo_carta':        TIPO_CARTA.get(r.tipo_carta, r.tipo_carta),
                'es_provisional':    r.es_provisional,
                'empleador': {
                    'nombre':  r.nombre_empleador,
                    'cedula':  r.cedula_empleador,
                    'usuario': _usuario_info(r.usuario_empleador) if r.usuario_empleador else None,
                },
                'empleado': {
                    'nombre':   c.nombre_completo,
                    'documento': c.documento_id,
                    'cargo':    c.cargo,
                    'sede':     c.sede.nombre if c.sede else None,
                },
                'contrato_id': c.id,
            })

        return Response({
            'total':    len(data),
            'registros': data,
        })


class UsuariosParaProvisionaView(APIView):
    """GET: lista de usuarios activos para seleccionar como provisional."""
    permission_classes = [IsAuthenticated]

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def get(self, request):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        users = (
            User.objects
            .filter(is_active=True)
            .exclude(pk=request.user.pk)
            .order_by('nombres', 'apellidos')
        )
        return Response([_usuario_info(u) for u in users])


class ReporteRegistroFirmaView(APIView):
    """GET: descarga Excel del historial de firma empleador, requiere rango de fechas."""
    permission_classes = [IsAuthenticated]

    @require_permission(['can_manage_firma_gh'], app_label='Usuarios')
    def get(self, request):
        fecha_desde = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta = request.query_params.get('fecha_hasta', '').strip()

        if not fecha_desde or not fecha_hasta:
            return Response(
                {'error': 'Los parámetros fecha_desde y fecha_hasta son requeridos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import pytz

        qs = (
            RegistroFirmaEmpleador.objects
            .filter(
                fecha_generacion__date__gte=fecha_desde,
                fecha_generacion__date__lte=fecha_hasta,
            )
            .select_related('contrato', 'contrato__sede', 'usuario_empleador')
            .order_by('-fecha_generacion')
        )

        H_FILL  = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
        H_FONT  = Font(color='FFFFFF', bold=True, size=11)
        H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        EVEN_FILL = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
        thin   = Side(style='thin', color='CBD5E1')
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        HEADERS = [
            ('Fecha generación',      20),
            ('Tipo carta',            16),
            ('Empleador (nombre)',     30),
            ('Empleador (C.C.)',       16),
            ('Es provisional',        14),
            ('Empleado (nombre)',      30),
            ('Empleado (documento)',   18),
            ('Cargo',                 26),
            ('Sede',                  22),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Registros Firma'
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 28

        for col, (h, w_col) in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = H_FILL
            cell.font = H_FONT
            cell.alignment = H_ALIGN
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = w_col

        TIPO_CARTA = {'NO_PRORROGA': 'Sin prórroga', 'PRORROGA': 'Prórroga', 'TERMINACION': 'Terminación'}
        tz_co = pytz.timezone('America/Bogota')

        for row_n, r in enumerate(qs, 2):
            c = r.contrato
            fecha_local = r.fecha_generacion.astimezone(tz_co).strftime('%d/%m/%Y %H:%M')
            vals = [
                fecha_local,
                TIPO_CARTA.get(r.tipo_carta, r.tipo_carta),
                r.nombre_empleador,
                r.cedula_empleador,
                'Sí' if r.es_provisional else 'No',
                c.nombre_completo,
                c.documento_id,
                c.cargo or '',
                c.sede.nombre if c.sede else '',
            ]
            fill = EVEN_FILL if row_n % 2 == 0 else None
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_n, column=col, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        # Hoja de metadatos
        ws_m = wb.create_sheet('Filtros')
        ws_m.column_dimensions['A'].width = 20
        ws_m.column_dimensions['B'].width = 25
        bold = Font(bold=True)
        for row_i, (k, v) in enumerate([
            ('Fecha desde', fecha_desde),
            ('Fecha hasta', fecha_hasta),
            ('Total registros', qs.count()),
        ], 1):
            ws_m.cell(row=row_i, column=1, value=k).font = bold
            ws_m.cell(row=row_i, column=2, value=v)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'registros_firma_{fecha_desde}_a_{fecha_hasta}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response
