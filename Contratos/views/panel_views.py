from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from ..models import Contrato, AsignacionCentro
from ..serializers import AsignacionCentroSerializer
from datetime import date


class EscanearSiesaView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from ..tasks import revisar_contratos_60_dias
        try:
            antes = Contrato.objects.count()
            revisar_contratos_60_dias()
            despues = Contrato.objects.count()
            nuevos = despues - antes
            return Response({
                'mensaje': f'Escaneo completado. {nuevos} contrato(s) nuevo(s) generado(s).',
                'nuevos': nuevos,
                'total': despues,
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PanelResumenView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Contrato.objects.all()
        from ..models import AsignacionCentro
        sedes_asignadas = AsignacionCentro.objects.filter(
            usuario=request.user, activo=True
        ).values_list('sede', flat=True)
        if sedes_asignadas.exists():
            qs = qs.filter(sede__in=sedes_asignadas)

        from django.db.models import Q
        qs_activos = qs.exclude(estado='CANCELADO')
        # Carta pendiente de firma = tiene pdf_carta_key pero no pdf_firmado_key,
        # independientemente del estado (incluye PENDIENTE_DECISION_DIRECTOR urgentes).
        pendiente_firma = qs_activos.filter(
            Q(pdf_carta_key__isnull=False) & ~Q(pdf_carta_key=''),
            Q(pdf_firmado_key__isnull=True) | Q(pdf_firmado_key=''),
        ).exclude(estado='FIRMADO').count()
        return Response({
            'total': qs_activos.count(),
            'pendiente_firma': pendiente_firma,
            'pendiente_decision': qs_activos.filter(estado='PENDIENTE_DECISION_DIRECTOR').count(),
            'firmados': qs_activos.filter(estado='FIRMADO').count(),
            'sin_canal': qs_activos.filter(estado='SIN_CANAL_CONTACTO').count(),
        })


class ContratacionesView(APIView):
    """Historial de contratos firmados agrupados por empleado."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = (
            Contrato.objects
            .filter(estado='FIRMADO')
            .select_related('sede')
            .prefetch_related('documentos_adicionales')
            .order_by('nombre_completo', '-fecha_firma')
        )

        from ..models import AsignacionCentro
        sedes_asignadas = AsignacionCentro.objects.filter(
            usuario=request.user, activo=True
        ).values_list('sede', flat=True)
        if sedes_asignadas.exists():
            qs = qs.filter(sede__in=sedes_asignadas)

        search = request.query_params.get('search', '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(Q(nombre_completo__icontains=search) | Q(documento_id__icontains=search))

        fecha_desde = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta = request.query_params.get('fecha_hasta', '').strip()
        if fecha_desde:
            qs = qs.filter(fecha_firma__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_firma__date__lte=fecha_hasta)

        empleados = {}
        for c in qs:
            key = c.documento_id
            if key not in empleados:
                empleados[key] = {
                    'documento_id':    c.documento_id,
                    'tipo_documento':  c.tipo_documento,
                    'nombre_completo': c.nombre_completo,
                    'contratos': [],
                }
            docs_adicionales = c.documentos_adicionales.count()
            total_docs = (1 if c.pdf_firmado_key else 0) + docs_adicionales
            empleados[key]['contratos'].append({
                'id':                 c.id,
                'tipo_carta':         c.tipo_carta,
                'cargo':              c.cargo,
                'fecha_finalizacion': str(c.fecha_finalizacion) if c.fecha_finalizacion else None,
                'fecha_firma':        c.fecha_firma.isoformat() if c.fecha_firma else None,
                'sede_nombre':        c.sede.nombre if c.sede else None,
                'sede_codigo':        c.sede.codigo if c.sede else None,
                'tiene_pdf_firmado':  bool(c.pdf_firmado_key),
                'total_documentos':   total_docs,
            })

        lista = list(empleados.values())
        # Agregar total_documentos por empleado para el badge
        for emp in lista:
            emp['total_documentos'] = sum(c['total_documentos'] for c in emp['contratos'])

        return Response({
            'empleados':        lista,
            'total_empleados':  len(lista),
            'total_contratos':  sum(len(e['contratos']) for e in lista),
        })


class ReporteContratacionesView(APIView):
    """Genera y devuelve un Excel con los contratos firmados según los filtros activos."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.http import HttpResponse
        import pytz

        qs = (
            Contrato.objects
            .filter(estado='FIRMADO')
            .select_related('sede')
            .prefetch_related('documentos_adicionales')
            .order_by('nombre_completo', '-fecha_firma')
        )

        sedes_asignadas = AsignacionCentro.objects.filter(
            usuario=request.user, activo=True
        ).values_list('sede', flat=True)
        if sedes_asignadas.exists():
            qs = qs.filter(sede__in=sedes_asignadas)

        search = request.query_params.get('search', '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(Q(nombre_completo__icontains=search) | Q(documento_id__icontains=search))

        fecha_desde = request.query_params.get('fecha_desde', '').strip()
        fecha_hasta = request.query_params.get('fecha_hasta', '').strip()
        if fecha_desde:
            qs = qs.filter(fecha_firma__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_firma__date__lte=fecha_hasta)

        # ── estilos ──────────────────────────────────────────────────
        H_FILL  = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
        H_FONT  = Font(color='FFFFFF', bold=True, size=11)
        H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        EVEN_FILL = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
        thin   = Side(style='thin', color='CBD5E1')
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        HEADERS = [
            ('Nombre completo',      35),
            ('Tipo doc.',            10),
            ('Documento',            18),
            ('Tipo carta',           16),
            ('Cargo',                28),
            ('Sede',                 22),
            ('Inicio contrato',      16),
            ('Fecha finalización',   18),
            ('Duración prórroga',    16),
            ('Fecha fin prórroga',   16),
            ('Fecha firma',          22),
            ('Email',                30),
            ('Celular',              15),
            ('Docs. adicionales',    14),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Contrataciones'
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 28

        for col, (h, w) in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill   = H_FILL
            cell.font   = H_FONT
            cell.alignment = H_ALIGN
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = w

        TIPO_CARTA   = {'NO_PRORROGA': 'Sin prórroga', 'PRORROGA': 'Prórroga', 'TERMINACION': 'Terminación'}
        DUR_LABELS   = dict(Contrato.DURACION_PRORROGA_CHOICES)
        tz_co        = pytz.timezone('America/Bogota')

        def fmt_date(d):
            return d.strftime('%d/%m/%Y') if d else ''

        def fmt_dt(dt):
            if not dt:
                return ''
            return dt.astimezone(tz_co).strftime('%d/%m/%Y %H:%M')

        for row_n, c in enumerate(qs, 2):
            docs_count = c.documentos_adicionales.count()
            vals = [
                c.nombre_completo,
                c.tipo_documento,
                c.documento_id,
                TIPO_CARTA.get(c.tipo_carta, c.tipo_carta),
                c.cargo or '',
                c.sede.nombre if c.sede else '',
                fmt_date(c.fecha_inicio_contrato),
                fmt_date(c.fecha_finalizacion),
                DUR_LABELS.get(c.duracion_prorroga, '') if c.duracion_prorroga else '',
                fmt_date(c.fecha_fin_prorroga),
                fmt_dt(c.fecha_firma),
                c.email or '',
                c.celular or '',
                docs_count,
            ]
            fill = EVEN_FILL if row_n % 2 == 0 else None
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_n, column=col, value=val)
                cell.border    = BORDER
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        # ── hoja de metadatos ────────────────────────────────────────
        ws_m = wb.create_sheet('Filtros')
        ws_m.column_dimensions['A'].width = 22
        ws_m.column_dimensions['B'].width = 28
        bold = Font(bold=True)
        for r, (k, v) in enumerate([
            ('Búsqueda', search or 'Ninguna'),
            ('Fecha desde', fecha_desde or 'Sin filtro'),
            ('Fecha hasta', fecha_hasta or 'Sin filtro'),
            ('Total registros', qs.count()),
        ], 1):
            ws_m.cell(row=r, column=1, value=k).font = bold
            ws_m.cell(row=r, column=2, value=v)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        if fecha_desde and fecha_hasta:
            nombre_archivo = f'contrataciones_{fecha_desde}_a_{fecha_hasta}.xlsx'
        else:
            from datetime import date as dt_date
            nombre_archivo = f'contrataciones_{dt_date.today().isoformat()}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response


class AsignacionesSedView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = AsignacionCentro.objects.select_related('usuario', 'sede').filter(activo=True)
        return Response(AsignacionCentroSerializer(qs, many=True).data)

    def post(self, request):
        serializer = AsignacionCentroSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk=None):
        try:
            asignacion = AsignacionCentro.objects.get(pk=pk)
        except AsignacionCentro.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        asignacion.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
